import openpyxl

'''class home_pageData:
    test_home_page_data = [{"Name": "Sushmitha", "Email": "sushmithasj777@gmail.com", "Pwd": "1234567"},
                           {"Name": "AKHIL", "Email": "akhil@gmail.com", "Pwd": "1234567"}]'''
excel = openpyxl.load_workbook(
    "C:/Users/ssj/OneDrive - Amadeus Workplace/Desktop/pythontestdata.xlsx")
sheet = excel.active
cell = sheet.cell(row=2, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "sushu"
print(sheet.max_row)
print(sheet.max_column)

for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value)

